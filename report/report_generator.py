import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # 新增 Font 等样式类
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date
from sqlalchemy import and_, or_
import os
import json
from urllib.parse import quote
from config import REPORT_DIR, FILES_DIR, APP_PUBLIC_BASE_URL, TENDER_FILES_URL_PREFIX
from utils.log import log
from utils.db import get_db, TenderProject, ProjectStatus

# 仅 site_name 含平台名、且数据库 region 为空时的默认行政区（与爬虫命名一致）
_PLATFORM_DEFAULT_REGION = {
    "宁波市阳光采购服务平台": "宁波市",
    "杭州市公共资源交易网": "杭州市",
    "嘉兴禾采联综合采购服务平台": "嘉兴市",
    "绍兴市阳光采购服务平台": "绍兴市",
    "湖州市绿色采购服务平台": "湖州市",
    "义乌市阳光招标采购平台": "义乌市",
    "丽水市阳光采购服务平台": "丽水市",
    "衢州市阳光交易服务平台": "衢州市",
}


class ReportGenerator:
    def __init__(self):
        self.db = next(get_db())
        self.report_date = datetime.now().strftime("%Y-%m-%d")
        self.report_path = os.path.join(REPORT_DIR, f"标书资质匹配报告_{self.report_date}.xlsx")

    def _extract_procurement_type(self, site_name):
        """从site_name中提取采购类型，默认为公开招标"""
        # 采购类型默认为公开招标
        return "公开招标"

    def _infer_region_from_site(self, site_name, raw_region):
        """当数据库 region 为空时，从 site_name（如「平台名-杭州市」）或平台默认值推断行政区。"""
        r = (raw_region or "").strip()
        if r:
            return r
        sn = (site_name or "").strip()
        if not sn:
            return ""
        if "本地上传" in sn:
            return "本地"
        if "-" in sn:
            suffix = sn.split("-", 1)[-1].strip()
            if suffix:
                return suffix
        for plat_name, default_reg in _PLATFORM_DEFAULT_REGION.items():
            if plat_name in sn:
                return default_reg
        return ""

    def _is_likely_file_download_url(self, url: str) -> bool:
        """判断是否为附件/API 下载地址（非浏览器公告详情页）。"""
        if not url:
            return False
        u = url.lower()
        markers = (
            "ztbattachdownloadaction",
            "epointwebbuilder/pages/webbuildermis/attach",
            "/api/file/download",
            "fileserver/api/download",
            "/files/myupfiles/",
            "getfilebytype",
            "ygcg.nbcqjy.org:8072/files",
            "attachdownload",
            "/api/file/get",
        )
        if any(m in u for m in markers):
            return True
        # 直接链到常见文档后缀且路径像文件存储
        if any(u.endswith(ext) for ext in (".doc", ".docx", ".pdf", ".zip", ".xls", ".xlsx")):
            if "/files/" in u or "download" in u:
                return True
        return False

    def _build_platform_detail_url(self, site_name: str, project_id: str) -> str:
        """按来源平台构造公告详情页 URL（project_id 为各平台唯一标识）。"""
        if not project_id:
            return ""
        sn = site_name or ""
        if "浙江省政府采购网" in sn:
            return f"https://zfcg.czt.zj.gov.cn/site/detail?parentId=600007&articleId={project_id}"
        if "杭州市公共资源交易网" in sn:
            return f"https://ggzy.hzctc.hangzhou.gov.cn/detail/{project_id}"
        if "绍兴市阳光采购" in sn:
            return f"https://ygcg.sxjypt.com/detail?bulletinId={project_id}"
        if "宁波市阳光采购" in sn:
            return f"https://ygcg.nbcqjy.org:8071/page/projectinfo/signup.html?PrjId={project_id}"
        return ""

    def _tender_file_public_url(self, file_path, public_file_base_url=None):
        """将本地 tender_files 下的已存在文件转为可点击 HTTP 路径（依赖网关或静态服务映射）。"""
        base = (public_file_base_url or APP_PUBLIC_BASE_URL or "").strip().rstrip("/")
        if not base or not file_path:
            return ""
        path = os.path.abspath(os.path.normpath(file_path))
        root = os.path.abspath(FILES_DIR)
        if not os.path.isfile(path):
            return ""
        try:
            rel = os.path.relpath(path, root)
        except ValueError:
            return ""
        rel_norm = rel.replace("\\", "/")
        if rel_norm.startswith("../") or rel_norm == "..":
            return ""
        enc = "/".join(quote(seg, safe="") for seg in rel_norm.split("/"))
        prefix = (TENDER_FILES_URL_PREFIX or "").strip().rstrip("/")
        if prefix and not prefix.startswith("/"):
            prefix = "/" + prefix
        return f"{base}{prefix}/{enc}" if prefix else f"{base}/{enc}"

    def _resolve_source_url(
            self,
            site_name,
            download_url,
            project_id,
            file_path=None,
            public_file_base_url=None,
    ):
        """报告「来源网站」列：宁波优先本机标书 HTTP 链接；其余逻辑同前。"""
        du = (download_url or "").strip()
        sn = site_name or ""
        if "本地上传" in sn and not du:
            return ""

        if "宁波市阳光采购" in sn:
            local_url = self._tender_file_public_url(file_path, public_file_base_url)
            if local_url:
                return local_url

        if du.startswith("http"):
            if self._is_likely_file_download_url(du):
                built = self._build_platform_detail_url(sn, project_id or "")
                return built or du
            return du

        built = self._build_platform_detail_url(sn, project_id or "")
        if built:
            return built
        return du or ""

    def _extract_province_city(self, region):
        """从region中提取省份和城市

        根据aaaa.py，API返回的districtName可能是：
        - 市级：杭州市、宁波市等
        - 区级：拱墅区、余杭区、临平区、上城区、滨江区等
        - 县级：淳安县、桐庐县等
        """
        if not region:
            return "未知", "未知"

        province = "浙江省"  # 默认省份

        # 精确匹配区域名称
        region = region.strip()  # 去除前后空格
        if region == "本地":
            return "本地", "本地"

        # 浙江省的城市和区县映射（根据aaaa.py返回的实际数据）
        city_mapping = {
            # 市级
            "浙江省本级": "浙江省本级",
            "杭州市": "杭州市",
            "宁波市": "宁波市",
            "温州市": "温州市",
            "嘉兴市": "嘉兴市",
            "湖州市": "湖州市",
            "绍兴市": "绍兴市",
            "金华市": "金华市",
            "衢州市": "衢州市",
            "舟山市": "舟山市",
            "台州市": "台州市",
            "丽水市": "丽水市",
            "义乌市": "义乌市",
            "本地": "本地",
            # 杭州市的区县（根据aaaa.py返回的数据）
            "拱墅区": "杭州市",
            "余杭区": "杭州市",
            "临平区": "杭州市",
            "上城区": "杭州市",
            "滨江区": "杭州市",
            "西湖区": "杭州市",
            "萧山区": "杭州市",
            "钱塘区": "杭州市",
            "富阳区": "杭州市",
            "临安区": "杭州市",
            "建德市": "杭州市",
            "淳安县": "杭州市",
            "桐庐县": "杭州市",
            # 宁波市区县
            "海曙区": "宁波市",
            "江北区": "宁波市",
            "北仑区": "宁波市",
            "镇海区": "宁波市",
            "鄞州区": "宁波市",
            "奉化区": "宁波市",
            "余姚市": "宁波市",
            "慈溪市": "宁波市",
            "宁海县": "宁波市",
            "象山县": "宁波市",
            # 温州区县（常见）
            "鹿城区": "温州市",
            "龙湾区": "温州市",
            "瓯海区": "温州市",
            "洞头区": "温州市",
            "瑞安市": "温州市",
            "乐清市": "温州市",
            "龙港市": "温州市",
            "永嘉县": "温州市",
            "平阳县": "温州市",
            "苍南县": "温州市",
            "文成县": "温州市",
            "泰顺县": "温州市",
            # 嘉兴区县
            "南湖区": "嘉兴市",
            "秀洲区": "嘉兴市",
            "嘉善县": "嘉兴市",
            "海盐县": "嘉兴市",
            "海宁市": "嘉兴市",
            "平湖市": "嘉兴市",
            "桐乡市": "嘉兴市",
            # 湖州区县
            "吴兴区": "湖州市",
            "南浔区": "湖州市",
            "德清县": "湖州市",
            "长兴县": "湖州市",
            "安吉县": "湖州市",
            # 绍兴区县
            "越城区": "绍兴市",
            "柯桥区": "绍兴市",
            "上虞区": "绍兴市",
            "诸暨市": "绍兴市",
            "嵊州市": "绍兴市",
            "新昌县": "绍兴市",
            # 金华区县
            "婺城区": "金华市",
            "金东区": "金华市",
            "兰溪市": "金华市",
            "东阳市": "金华市",
            "永康市": "金华市",
            "武义县": "金华市",
            "浦江县": "金华市",
            "磐安县": "金华市",
            # 衢州区县
            "柯城区": "衢州市",
            "衢江区": "衢州市",
            "江山市": "衢州市",
            "常山县": "衢州市",
            "开化县": "衢州市",
            "龙游县": "衢州市",
            # 舟山区县
            "定海区": "舟山市",
            "普陀区": "舟山市",
            "岱山县": "舟山市",
            "嵊泗县": "舟山市",
            # 台州区县
            "椒江区": "台州市",
            "黄岩区": "台州市",
            "路桥区": "台州市",
            "三门县": "台州市",
            "天台县": "台州市",
            "仙居县": "台州市",
            "温岭市": "台州市",
            "临海市": "台州市",
            "玉环市": "台州市",
            # 丽水区县
            "莲都区": "丽水市",
            "青田县": "丽水市",
            "缙云县": "丽水市",
            "遂昌县": "丽水市",
            "松阳县": "丽水市",
            "云和县": "丽水市",
            "庆元县": "丽水市",
            "景宁畲族自治县": "丽水市",
            "龙泉市": "丽水市",
            # 其他可能的区县
            "无区域": "未知"
        }

        # 优先精确匹配
        if region in city_mapping:
            city = city_mapping[region]
        else:
            # 如果精确匹配失败，尝试部分匹配
            # 如果包含"区"或"县"，可能是区县，需要映射到对应的市
            if "区" in region or "县" in region:
                # 尝试匹配市级名称
                for city_name in ["杭州市", "宁波市", "温州市", "嘉兴市", "湖州市", "绍兴市",
                                  "金华市", "衢州市", "舟山市", "台州市", "丽水市"]:
                    if city_name in region or region in city_name:
                        city = city_name
                        break
                else:
                    # 如果没匹配到，默认归到杭州市（因为大部分区县都在杭州市）
                    city = "杭州市"
            elif "市" in region:
                # 直接是市级名称
                city = region if region.endswith("市") else region + "市"
            else:
                # 尝试部分匹配
                city = "未知"
                for key, value in city_mapping.items():
                    if key in region or region in key:
                        city = value
                        break

        return province, city

    def _extract_objective_attainable_score(self, comparison_result):
        """从comparison_result中提取客观分可得分

        Args:
            comparison_result: 资质比对结果（文本字符串）

        Returns:
            客观分可得分（浮点数），如果无法提取则返回None
        """
        if not comparison_result:
            return None

        try:
            import re

            # 尝试多种格式匹配"客观分可得分"
            attainable_patterns = [
                r'客观分可得分[：:]\s*(\d+(?:\.\d+)?)\s*分',
                r'客观分可得分[：:]\s*\[(\d+(?:\.\d+)?)\]\s*分',
                r'客观分可得分\s*[：:]\s*(\d+(?:\.\d+)?)\s*分',
                r'可得分[：:]\s*(\d+(?:\.\d+)?)\s*分',
                r'可得分[：:]\s*\[(\d+(?:\.\d+)?)\]\s*分',
                r'可得分\s*[：:]\s*(\d+(?:\.\d+)?)\s*分',
            ]

            # 尝试匹配可得分
            for pattern in attainable_patterns:
                match = re.search(pattern, comparison_result)
                if match:
                    try:
                        attainable_score = float(match.group(1))
                        return attainable_score
                    except ValueError:
                        continue

            return None
        except Exception as e:
            log.debug(f"从comparison_result中提取客观分可得分失败：{str(e)}")
            return None

    def _get_project_data(
            self,
            start_date=None,
            end_date=None,
            regions=None,
            procurement_types=None,
            platform_code=None,
            public_file_base_url=None,
    ):
        """获取项目数据（支持筛选）

        Args:
            start_date: 开始日期（datetime或date对象）
            end_date: 结束日期（datetime或date对象）
            regions: 城市列表（如["杭州市", "宁波市"]），会按提取的城市进行筛选，None表示全选
            procurement_types: 采购类型列表（如["政府采购", "国企采购"]），None表示全选
            platform_code: 平台代码（如"zhejiang"、"hangzhou"），None表示全选
            public_file_base_url: 报告内文件外链根地址（如 https://a.com），覆盖 config.APP_PUBLIC_BASE_URL
        """
        query = self.db.query(TenderProject)

        # 时间筛选
        if start_date:
            if isinstance(start_date, date) and not isinstance(start_date, datetime):
                start_date = datetime.combine(start_date, datetime.min.time())
            query = query.filter(TenderProject.publish_time >= start_date)

        if end_date:
            if isinstance(end_date, date) and not isinstance(end_date, datetime):
                end_date = datetime.combine(end_date, datetime.max.time())
            query = query.filter(TenderProject.publish_time <= end_date)

        projects = query.all()
        data = []
        for proj in projects:
            # 提取采购类型
            procurement_type = self._extract_procurement_type(proj.site_name)

            region_text = self._infer_region_from_site(proj.site_name, getattr(proj, "region", None))
            if proj.region or region_text:
                log.debug(f"项目 {proj.id} region 原始={proj.region!r} 推断={region_text!r}")

            province, city = self._extract_province_city(region_text)

            # 城市筛选（按提取的城市进行筛选）
            if regions and len(regions) > 0:
                if city not in regions and region_text not in regions:
                    continue

            # 采购类型筛选
            if procurement_types and len(procurement_types) > 0:
                if procurement_type not in procurement_types:
                    continue

            # 平台筛选
            if platform_code:
                # 从site_name中提取平台代码（避免循环导入）
                site_name = proj.site_name or ""
                project_platform = None
                platform_map = {
                    "浙江省政府采购网": "zhejiang",
                    "杭州市公共资源交易网": "hangzhou",
                    "嘉兴禾采联综合采购服务平台": "jiaxing",
                    "宁波市阳光采购服务平台": "ningbo",
                    "绍兴市阳光采购服务平台": "shaoxing",
                    "湖州市绿色采购服务平台": "huzhou",
                    "义乌市阳光招标采购平台": "yiwu",
                    "丽水市阳光采购服务平台": "lishui",
                    "衢州市阳光交易服务平台": "quzhou",
                }
                for platform_name, code in platform_map.items():
                    if platform_name in site_name:
                        project_platform = code
                        break

                if project_platform != platform_code:
                    continue

            final_decision = proj.final_decision if hasattr(proj, 'final_decision') and proj.final_decision else "未判定"

            # 从comparison_result中提取客观分可得分
            objective_attainable_score = self._extract_objective_attainable_score(proj.comparison_result)
            objective_attainable_score_str = f"{objective_attainable_score:.1f}" if objective_attainable_score is not None else ""

            # 使用发布时间（publish_time），已经正常提取，无需异常检查
            # 时间格式：只保留日期，舍去小时、分钟和秒
            if proj.publish_time:
                publish_time_str = proj.publish_time.strftime("%Y-%m-%d")
            else:
                # 如果发布时间为空，使用创建时间作为备选（极少情况）
                if hasattr(proj, 'create_time') and proj.create_time:
                    publish_time_str = proj.create_time.strftime("%Y-%m-%d")
                else:
                    publish_time_str = "未知"

            source_url = self._resolve_source_url(
                proj.site_name,
                proj.download_url,
                getattr(proj, "project_id", None),
                file_path=getattr(proj, "file_path", None),
                public_file_base_url=public_file_base_url,
            )

            data.append({
                "项目ID": proj.id,
                "项目名称": proj.project_name,
                "省份": province,
                "城市": city,
                "区域": region_text or "未知",
                "采购类型": procurement_type,
                "来源网站": source_url,
                "发布时间": publish_time_str,
                "文件格式": proj.file_format or "",
                "状态": proj.status.value,
                "最终判定": final_decision,
                "客观分总分值": objective_attainable_score_str,
                "错误信息": proj.error_msg or "无"
            })
        return pd.DataFrame(data)

    def _get_qualified_projects(self, all_data):
        """筛选可参与项目

        说明：
        - 系统内部可能使用多种文案表示“推荐参与”，例如：
          - “可以参与”（早期版本）
          - “客观分满分”（基于评分逻辑的判定说明）
          - “推荐参与”（当前 AI 分析 run 方法使用的文案）
        - 这里统一视为“可参与”，后续如需新增文案，只需补充列表。
        """
        qualified_flags = ["可以参与", "客观分满分", "推荐参与"]
        return all_data[all_data["最终判定"].isin(qualified_flags)].copy()

    def _add_style_to_workbook(self, wb):
        """为Excel添加样式和自动筛选"""
        # 表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        # 对齐样式
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        # 边框样式
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 设置列宽（根据实际列数动态调整）
            column_widths = {
                'A': 10,  # 项目ID
                'B': 40,  # 项目名称
                'C': 12,  # 省份
                'D': 12,  # 城市
                'E': 15,  # 区域
                'F': 12,  # 采购类型
                'G': 25,  # 来源网站
                'H': 18,  # 发布时间
                'I': 10,  # 文件格式
                'J': 10,  # 状态
                'K': 12,  # 最终判定
                'L': 15,  # 客观分总分值
                'M': 30,  # 错误信息
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # 为表头添加样式
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = border

                # 为数据行添加边框和对齐
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for idx, cell in enumerate(row):
                        cell.border = border
                        # 项目名称左对齐，其他居中
                        if idx == 1:  # 项目名称列
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_center

                # 添加自动筛选
                if ws.max_row > 1:
                    ws.auto_filter.ref = ws.dimensions

        return wb

    def generate_report(
            self,
            start_date=None,
            end_date=None,
            regions=None,
            procurement_types=None,
            platform_code=None,
            report_filename=None,
            public_file_base_url=None,
    ):
        """生成Excel报告（支持筛选）

        Args:
            start_date: 开始日期（datetime或date对象），None表示不限制
            end_date: 结束日期（datetime或date对象），None表示不限制
            regions: 区域列表（如["杭州市", "宁波市"]），None或空列表表示全选
            procurement_types: 采购类型列表（如["政府采购", "国企采购"]），None或空列表表示全选
            platform_code: 平台代码（如"zhejiang"、"hangzhou"），None表示全选
            report_filename: 报告文件名，None表示使用默认名称
            public_file_base_url: 宁波等项目「来源网站」文件外链根地址，None 则用 config 或仅官方页
        """
        try:
            log.info(f"开始生成报告：{self.report_date}")
            if start_date:
                log.info(f"时间范围：{start_date} 至 {end_date}")
            if regions and len(regions) > 0:
                log.info(f"筛选区域：{regions}")
            if procurement_types and len(procurement_types) > 0:
                log.info(f"筛选采购类型：{procurement_types}")
            if platform_code:
                log.info(f"筛选平台：{platform_code}")

            # 1. 获取数据（应用筛选条件）
            all_data = self._get_project_data(
                start_date=start_date,
                end_date=end_date,
                regions=regions,
                procurement_types=procurement_types,
                platform_code=platform_code,
                public_file_base_url=public_file_base_url,
            )

            if len(all_data) == 0:
                raise ValueError("没有符合筛选条件的项目数据")

            # 2. 创建Excel工作簿
            wb = Workbook()
            # 删除默认工作表
            wb.remove(wb.active)

            # 3. 添加「项目列表」工作表
            ws = wb.create_sheet("项目列表")
            for r in dataframe_to_rows(all_data, index=False, header=True):
                ws.append(r)

            # 4. 添加样式和自动筛选
            wb = self._add_style_to_workbook(wb)

            # 5. 保存报告
            os.makedirs(REPORT_DIR, exist_ok=True)
            if report_filename:
                report_path = os.path.join(REPORT_DIR, report_filename)
            else:
                # 生成带时间戳的文件名
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report_path = os.path.join(REPORT_DIR, f"标书资质匹配报告_{timestamp}.xlsx")

            wb.save(report_path)
            log.info(f"报告生成完成：{report_path}，共{len(all_data)}条记录")
            return report_path
        except Exception as e:
            log.error(f"报告生成失败：{str(e)}")
            raise


if __name__ == "__main__":
    generator = ReportGenerator()
    generator.generate_report()